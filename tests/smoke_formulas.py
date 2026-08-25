"""冒烟测试：openpyxl 写公式 -> formulas 重算 -> 读回断言。
这条链路是全项目勾稽验证的地基，先证明它能跑。"""
import openpyxl, formulas, os

os.chdir(os.path.dirname(__file__))
F = "smoke.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "T"
ws["A1"], ws["B1"], ws["C1"] = "主体", "期间", "值"
rows = [("云", "2023", 100), ("云", "2024", 200), ("行业", "2023", 50)]
for i, (a, b, c) in enumerate(rows, 2):
    ws[f"A{i}"], ws[f"B{i}"], ws[f"C{i}"] = a, b, c

# 需要验证的函数族：SUMIFS(含"全部"的<>通配) / INDEX+MATCH / IF / 四则 / 跨行SUM
ws["E1"] = "云"
ws["E2"] = '=SUMIFS(C2:C4,A2:A4,E1,B2:B4,"2024")'          # 期望 200
ws["E3"] = '=SUMIFS(C2:C4,A2:A4,"<>",B2:B4,"2023")'          # 期望 150（"全部"用法）
ws["E4"] = '=INDEX(C2:C4,MATCH("行业",A2:A4,0))'              # 期望 50
ws["E5"] = '=IF(E2>150,"BIG","SMALL")'                        # 期望 BIG
ws["E6"] = "=SUM(C2:C4)*1.05-10.5"                            # 期望 355.0
ws["E7"] = "=E2/0"                                            # 期望 #DIV/0!（错误也能读）
wb.save(F)

xl = formulas.ExcelModel().loads(F).finish()
sol = xl.calculate()

def read(cell):
    for k, v in sol.items():
        if k.upper().endswith(f"]T'!{cell}") or k.upper().endswith(f"]T'!{cell}[0,0]"):
            try:
                return v.value[0, 0]
            except Exception:
                return v.value
    raise KeyError(cell)

checks = {
    "E2": 200, "E3": 150, "E4": 50, "E5": "BIG", "E6": 355.0,
}
fails = []
for c, exp in checks.items():
    got = read(c)
    ok = abs(got - exp) < 1e-9 if isinstance(exp, (int, float)) else got == exp
    print(f"{c}: 期望={exp!r} 实得={got!r} {'✅' if ok else '❌'}")
    if not ok:
        fails.append(c)
print("E7(错误值):", read("E7"))
print("\n结论:", "地基可用 ✅" if not fails else f"地基不稳 ❌ {fails}")
