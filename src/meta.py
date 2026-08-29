"""Meta 宽表数据层 —— 数据与逻辑分离的落点。

数据流：行业配置(assumptions) ─初始化→ Meta 宽表(rows) ─查询→ 引擎/Excel
进公司后：ERP/BI 导出 ─按期灌入→ 同一张 Meta 宽表，引擎与公式零改动。

本模块同时暴露与旧 assumptions 同名的兼容层，engine.py 只需
`import meta as A` 即可整体切换数据源——引擎逻辑与数据源解耦的可运行证明。
"""
import assumptions as _cfg          # 行业配置（大模型公司实例）
from metric_dict import METRICS

YEARS = _cfg.YEARS
N = _cfg.N
ENTITIES = _cfg.ENTITIES
ENTITY_NAMES = _cfg.ENTITY_NAMES
SEGMENTS = _cfg.SEGMENTS
REGIONS = _cfg.REGIONS

def _scenario(year):
    return "ACTUAL" if year <= 2025 else "BUDGET"

def build_rows(cfg=_cfg):
    """从行业配置生成 Meta 宽表 rows。每行一个 metric×entity×region×period。"""
    rows = []
    def add(metric, value, period, entity="GROUP", region=None, source="假设初始化", note=""):
        rows.append({"entity": entity, "segment": None, "region": region,
                     "period": YEARS[period], "scenario": _scenario(YEARS[period]),
                     "metric": metric, "value": float(value), "source": source, "note": note})
    for i in range(N):
        # 集团
        add("tax_rate", cfg.TAX_RATE, i); add("dso_days", cfg.DSO, i)
        add("dpo_days", cfg.DPO, i); add("ic_days", cfg.IC_DAYS, i)
        add("wacc", cfg.WACC, i)
        for rg in cfg.REGIONS:
            add("region_mix", cfg.REGION_MIX[rg][i], i, region=rg,
                source="行业研究", note="收入地区占比")
        # 云
        add("token_volume_day", cfg.CLOUD_TOKENS_B[i], i, "CLOUD")
        add("api_price", cfg.CLOUD_PRICE[i], i, "CLOUD")
        add("inference_ucost", cfg.CLOUD_UCOST[i], i, "CLOUD")
        add("bandwidth_fee", cfg.CLOUD_BW[i], i, "CLOUD")
        add("customers", cfg.CLOUD_CUSTOMERS[i], i, "CLOUD")
        add("churn_m", cfg.CLOUD_CHURN_M, i, "CLOUD")
        # 行业
        add("projects_pvt", cfg.IND_PROJECTS[i], i, "IND")
        add("avg_contract", cfg.IND_CONTRACT[i], i, "IND")
        add("subscribers_sol", cfg.IND_SUBSCRIBERS[i], i, "IND")
        add("sub_fee", cfg.IND_SUB_FEE[i], i, "IND")
        add("cogs_rate_pvt", cfg.IND_COGS_PVT, i, "IND")
        add("cogs_rate_sol", cfg.IND_COGS_SOL, i, "IND")
        # 互联
        add("mau_w", cfg.TOC_MAU_W[i], i, "TOC")
        add("pay_rate", cfg.TOC_PAYRATE[i], i, "TOC")
        add("arppu", cfg.TOC_ARPPU[i], i, "TOC")
        add("ads_rev", cfg.TOC_ADS[i], i, "TOC")
        add("ucost_toc", cfg.TOC_UCOST[i], i, "TOC")
        add("cac", cfg.TOC_CAC[i], i, "TOC", source="投放实测")
        add("new_user_mult", cfg.TOC_NEW_U_MULTIPLIER, i, "TOC")
        # 编制×主体
        for ent, key in [("CLOUD", "CLOUD"), ("IND", "IND"), ("TOC", "TOC")]:
            for role in ["rd", "sales", "gna"]:
                add(f"headcount_{role}", cfg.__dict__[f"{key}_HEADCOUNT"][role][i], i, ent)
                add(f"salary_{role}", cfg.__dict__[f"{key}_SALARY"][role], i, ent)
        add("headcount_rd", cfg.HOLD_HEADCOUNT[i], i, "HOLD")
        add("headcount_gna", cfg.HOLD_GNA_HC[i], i, "HOLD")
        add("salary_gna", cfg.HOLD_GNA_SALARY, i, "HOLD")
        add("other_rd", cfg.HOLD_OTHER_RD[i], i, "HOLD")
        add("hold_salary", cfg.HOLD_SALARY[i], i, "HOLD")
        # 关联定价
        for ent in ["CLOUD", "IND", "TOC"]:
            add("t1_rate", cfg.T1_RATE[ent], i, ent)
        add("t2_fee", cfg.T2_FEE[i], i, "HOLD")
        add("t3_share", cfg.T3_SHARE, i); add("t3_gm", cfg.T3_GM, i)
        add("cutoff_2027", cfg.CUTOFF_2027, i)
        # 融资/期初
        for ent in cfg.ENTITIES:
            add("equity_inject", cfg.EQUITY_INJECT[ent][i], i, ent)
        for ent in cfg.ENTITIES:
            add("open_re", cfg.OPENING[ent]["RE"], i, ent, note="仅2023期初有效")
            add("open_pi", cfg.OPENING[ent]["paid_in"], i, ent, note="仅2023期初有效")
        add("open_fa", cfg.OPENING_FA["HOLD"], i, "HOLD", note="训练卡")
        add("open_fa", cfg.OPENING_FA["CLOUD"], i, "CLOUD", note="推理卡")
        add("open_lt_inv", cfg.OPENING_LT_INV, i, "HOLD")
        for ent in ["CLOUD", "IND", "TOC"]:
            add("deferred_rate", cfg.DEFERRED_RATE[ent], i, ent)
        add("capex_gpu", cfg.CAPEX["HOLD_TRAIN"][i], i, "HOLD", note="训练卡")
        add("capex_gpu", cfg.CAPEX["CLOUD_INFER"][i], i, "CLOUD", note="推理卡")
    return rows

class _Index:
    def __init__(self, rows):
        self.d = {(r["metric"], r["entity"], r["region"], r["period"]): r["value"] for r in rows}
    def get(self, metric, i, entity="GROUP", region=None):
        key = (metric, entity, region, YEARS[i])
        if key not in self.d:
            raise KeyError(f"Meta 缺行: {key}")
        return self.d[key]

_IDX = _Index(build_rows())

def get(metric, i, entity="GROUP", region=None):
    """引擎统一取数接口：get(指标代码, 期index, 主体, 地区)。"""
    return _IDX.get(metric, i, entity, region)

def bind(rows):
    """切换数据源（如 load_meta_from_xlsx 的返回值）——引擎零改动。"""
    global _IDX
    _IDX = _Index(rows)

def load_meta_from_xlsx(path):
    """进公司后的灌数入口：从任意含 Meta sheet 的 xlsx 读回宽表。"""
    from openpyxl import load_workbook
    ws = load_workbook(path)["Meta"]
    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r[3] or not r[5]:
            continue
        rows.append({"entity": r[0], "segment": r[1], "region": r[2], "period": int(r[3]),
                     "scenario": r[4], "metric": r[5], "value": float(r[6]),
                     "source": r[7], "note": r[8] or ""})
    return rows

def rows():
    return build_rows()

# ─────────────────────────────────────────────────────
# 兼容层：与 assumptions 同名常量，值全部来自 Meta 查询
# （engine.py / 生成器 import meta as A 即无缝切换）
# ─────────────────────────────────────────────────────
TAX_RATE = get("tax_rate", 0); DSO = int(get("dso_days", 0))
DPO = int(get("dpo_days", 0)); IC_DAYS = int(get("ic_days", 0)); WACC = get("wacc", 0)
CLOUD_TOKENS_B = [get("token_volume_day", i, "CLOUD") for i in range(N)]
CLOUD_PRICE = [get("api_price", i, "CLOUD") for i in range(N)]
CLOUD_UCOST = [get("inference_ucost", i, "CLOUD") for i in range(N)]
CLOUD_BW = [get("bandwidth_fee", i, "CLOUD") for i in range(N)]
CLOUD_CUSTOMERS = [get("customers", i, "CLOUD") for i in range(N)]
CLOUD_CHURN_M = get("churn_m", 0, "CLOUD")
CLOUD_HEADCOUNT = {k: [get(f"headcount_{k}", i, "CLOUD") for i in range(N)] for k in ["rd", "sales", "gna"]}
CLOUD_SALARY = {k: get(f"salary_{k}", 0, "CLOUD") for k in ["rd", "sales", "gna"]}
IND_PROJECTS = [get("projects_pvt", i, "IND") for i in range(N)]
IND_CONTRACT = [get("avg_contract", i, "IND") for i in range(N)]
IND_SUBSCRIBERS = [get("subscribers_sol", i, "IND") for i in range(N)]
IND_SUB_FEE = [get("sub_fee", i, "IND") for i in range(N)]
IND_COGS_PVT = get("cogs_rate_pvt", 0, "IND"); IND_COGS_SOL = get("cogs_rate_sol", 0, "IND")
IND_HEADCOUNT = {k: [get(f"headcount_{k}", i, "IND") for i in range(N)] for k in ["rd", "sales", "gna"]}
IND_SALARY = {k: get(f"salary_{k}", 0, "IND") for k in ["rd", "sales", "gna"]}
TOC_MAU_W = [get("mau_w", i, "TOC") for i in range(N)]
TOC_PAYRATE = [get("pay_rate", i, "TOC") for i in range(N)]
TOC_ARPPU = [get("arppu", i, "TOC") for i in range(N)]
TOC_ADS = [get("ads_rev", i, "TOC") for i in range(N)]
TOC_UCOST = [get("ucost_toc", i, "TOC") for i in range(N)]
TOC_CAC = [get("cac", i, "TOC") for i in range(N)]
TOC_NEW_U_MULTIPLIER = get("new_user_mult", 0, "TOC")
TOC_HEADCOUNT = {k: [get(f"headcount_{k}", i, "TOC") for i in range(N)] for k in ["rd", "sales", "gna"]}
TOC_SALARY = {k: get(f"salary_{k}", 0, "TOC") for k in ["rd", "sales", "gna"]}
HOLD_HEADCOUNT = [get("headcount_rd", i, "HOLD") for i in range(N)]
HOLD_SALARY = [get("hold_salary", i, "HOLD") for i in range(N)]
HOLD_OTHER_RD = [get("other_rd", i, "HOLD") for i in range(N)]
HOLD_GNA_HC = [get("headcount_gna", i, "HOLD") for i in range(N)]
HOLD_GNA_SALARY = get("salary_gna", 0, "HOLD")
HOLD_INITIAL_INVESTMENT = _cfg.HOLD_INITIAL_INVESTMENT
T1_RATE = {e: get("t1_rate", 0, e) for e in ["CLOUD", "IND", "TOC"]}
T2_FEE = [get("t2_fee", i, "HOLD") for i in range(N)]
T3_SHARE = get("t3_share", 0); T3_GM = get("t3_gm", 0)
CUTOFF_2027 = get("cutoff_2027", 0)
EQUITY_INJECT = {e: [get("equity_inject", i, e) for i in range(N)] for e in ENTITIES}
REGION_MIX = {rg: [get("region_mix", i, region=rg) for i in range(N)] for rg in REGIONS}
CAPEX = {"HOLD_TRAIN": [get("capex_gpu", i, "HOLD") for i in range(N)],
         "CLOUD_INFER": [get("capex_gpu", i, "CLOUD") for i in range(N)]}
CAPEX_OPENING = {"HOLD_TRAIN": get("open_fa", 0, "HOLD"), "CLOUD_INFER": get("open_fa", 0, "CLOUD")}
DEP_LIFE = _cfg.DEP_LIFE
OPENING = {e: {"RE": get("open_re", 0, e), "paid_in": get("open_pi", 0, e)} for e in ENTITIES}
OPENING_FA = {"HOLD": get("open_fa", 0, "HOLD"), "CLOUD": get("open_fa", 0, "CLOUD")}
OPENING_LT_INV = get("open_lt_inv", 0, "HOLD")
DEFERRED_RATE = {e: get("deferred_rate", 0, e) for e in ["CLOUD", "IND", "TOC"]}

if __name__ == "__main__":
    print(f"Meta 宽表：{len(rows())} 行 | 指标字典：{len(METRICS)} 个")
    print("冒烟：CLOUD_TOKENS_B =", CLOUD_TOKENS_B)
    print("冒烟：REGION_MIX 华东 =", REGION_MIX["华东"])
