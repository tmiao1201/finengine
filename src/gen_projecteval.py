"""私有化项目评估 ProjectEval：3 候选项目现金流 → NPV@WACC / IRR → P1 敏感性矩阵 → 决策结论。
口径：Year0 恒 0（无前期垫资，Y1 起算）；交付期收入=合同额/交付年数、交付成本=合同额×当年成本比例；
交付完成后每年净现金流=运维收入×(1−运维成本率)。NPV=NPV(WACC, Y1:Y5)——Excel NPV 语义首参按
第 1 期折现，故区间不含 Y0 列；IRR=IRR(Y0:Y5) 全列。formulas 库实测（/tmp 冒烟）：NPV 支持
（含"率+区间"与"率+内联表达式"两形态）、IRR 支持（前导 0 无碍、无符号变化返回 #NUM!）、
IFERROR 支持——全程净流入为正的项目 IRR 无有限解，表内以 IFERROR 显示「—」。单位：万元。"""
import math
import assumptions as A
from style import *
from layout import Layout
from openpyxl.utils import get_column_letter
import numpy_financial as npf

SHEET = "ProjectEval"
NYRS = 6                            # Year0-5 共 6 列（C-H）
MULTS = [0.80, 0.90, 1.10, 1.20]    # ±20% 四档（敏感性矩阵行/列各 4 档）

# 3 个候选私有化项目（虚构，量级贴合行业子公司平均合同额 1000-1500 万）
PROJECTS = [
    {"id": "P1", "name": "华北某银行 · 风控大模型", "contract": 1800.0, "dyears": 2,
     "cost_ratio": [0.55, 0.25], "maint": 180.0, "mcr": 0.40},
    {"id": "P2", "name": "华东某车企 · 智能座舱", "contract": 1200.0, "dyears": 1,
     "cost_ratio": [0.65], "maint": 120.0, "mcr": 0.35},
    {"id": "P3", "name": "华南某政务 · 热线助手", "contract": 900.0, "dyears": 1,
     "cost_ratio": [0.60], "maint": 150.0, "mcr": 0.30},
]

# ── 影子引擎（期望值，供逐格交叉验证）─────────────────────
def _rev_flows(p):
    return [0.0] + [p["contract"] / p["dyears"]] * p["dyears"] + [p["maint"]] * (5 - p["dyears"])

def _cost_flows(p):
    return ([0.0] + [p["contract"] * x for x in p["cost_ratio"]]
            + [p["maint"] * p["mcr"]] * (5 - p["dyears"]))

def _flows(p):
    return [r - c for r, c in zip(_rev_flows(p), _cost_flows(p))]

def _npv(cf, w=A.WACC):
    return sum(cf[i] / (1 + w) ** i for i in range(1, 6))

def _irr_or_dash(cf):
    try:
        v = float(npf.irr(cf))
        return v if math.isfinite(v) else "—"
    except Exception:
        return "—"

def _sens_npv(p, m, k, w=A.WACC):
    """敏感性：收入=合同额×m；交付成本=合同额×成本比例×k；运维现金流不变。"""
    d = p["dyears"]
    fl = [p["contract"] * m / d - p["contract"] * p["cost_ratio"][j] * k for j in range(d)]
    fl += [p["maint"] * (1 - p["mcr"])] * (5 - d)
    return sum(v / (1 + w) ** (i + 1) for i, v in enumerate(fl))


def gen(wb, L, eng):
    ws = wb.create_sheet(SHEET)
    ws.sheet_properties.tabColor = TAB["analysis"]
    ws["A1"] = "私有化项目评估 · ProjectEval"
    ws["A2"] = "智擎行业 · 3 候选项目：现金流 → NPV@WACC / IRR → P1 敏感性 → 决策 | 蓝字=输入 | 万元"
    ws["A1"].font = TITLE; ws["A2"].font = SUBTITLE
    ws.cell(row=3, column=1, value="科目（万元）").font = BOLD
    ws.cell(row=3, column=2, value="备注").font = BOLD
    for i in range(NYRS):
        c = ws.cell(row=3, column=3 + i, value=f"Y{i}")
        c.font = BOLD_WHITE; c.fill = HEADER_FILL; c.alignment = Alignment("center")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 34
    for i in range(NYRS):
        ws.column_dimensions[get_column_letter(3 + i)].width = 12.5
    ws.freeze_panes = "C4"

    r = 4
    def sec(t):
        nonlocal r
        c = ws.cell(row=r, column=1, value=t); c.font = BOLD; c.fill = SECTION_FILL
        for col in range(2, 9):
            ws.cell(row=r, column=col).fill = SECTION_FILL
        r += 1
    def label_cell(row, label, note=None, bold=False):
        c = ws.cell(row=row, column=1, value=label); c.font = BOLD if bold else BLACK_FORMULA
        if note:
            n = ws.cell(row=row, column=2, value=note); n.font = SUBTITLE
    def put(row, col, val, font=BLACK_FORMULA, numFmt=NUM, fill=None, bold=False):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font; cell.number_format = numFmt
        if fill: cell.fill = fill
        if bold: cell.border = TOP_BORDER
        return cell

    # ── 一、输入区（每项目一张卡片，全部 L.reg）────────────
    sec("── 一、项目输入（蓝字=可调，改动联动现金流 / 指标 / 敏感性）──")
    prows = {}
    for p in PROJECTS:
        c = ws.cell(row=r, column=1, value=f"{p['id']} · {p['name']}"); c.font = BOLD
        for col in range(1, 9):
            ws.cell(row=r, column=col).fill = INPUT_FILL
        r += 1
        rows = {}
        rows["contract"] = r
        label_cell(r, "合同额", "私有化一次性合同")
        put(r, 3, p["contract"], font=BLUE_INPUT)
        L.reg(SHEET, f"{p['id']}_contract", r, expected=(p["contract"],)); r += 1
        rows["dyears"] = r
        label_cell(r, "交付年数（年）", "交付期收入按年数均摊")
        put(r, 3, p["dyears"], font=BLUE_INPUT, numFmt='0')
        L.reg(SHEET, f"{p['id']}_dyears", r, expected=(p["dyears"],)); r += 1
        rows["cost"] = r
        label_cell(r, "交付成本比例（按年，占合同额）", "C列=Y1、D列=Y2（P1 两年交付）")
        for j, ratio in enumerate(p["cost_ratio"]):
            put(r, 3 + j, ratio, font=BLUE_INPUT, numFmt=PCT)
        L.reg(SHEET, f"{p['id']}_cost", r, expected=tuple(p["cost_ratio"])); r += 1
        rows["maint"] = r
        label_cell(r, "运维收入（万/年，交付完成后）")
        put(r, 3, p["maint"], font=BLUE_INPUT)
        L.reg(SHEET, f"{p['id']}_maint", r, expected=(p["maint"],)); r += 1
        rows["mcr"] = r
        label_cell(r, "运维成本率")
        put(r, 3, p["mcr"], font=BLUE_INPUT, numFmt=PCT)
        L.reg(SHEET, f"{p['id']}_mcr", r, expected=(p["mcr"],)); r += 1
        prows[p["id"]] = rows
    r += 1

    # ── 二、现金流区（每项目 收入/成本/净现金流 × Y0-5）────
    sec("── 二、净现金流（Y0 恒 0；交付期收入=合同额/交付年数、成本=合同额×当年成本比例；"
        "交付完成后净额=运维收入×(1−成本率)）──")
    for p in PROJECTS:
        rows = prows[p["id"]]
        cr, dr, cor = f"$C${rows['contract']}", f"$C${rows['dyears']}", rows["cost"]
        mr, mcr_r = f"$C${rows['maint']}", f"$C${rows['mcr']}"
        # 收入行
        label_cell(r, f"{p['id']} 收入（交付+运维）", "交付期=合同额/交付年数")
        rev_f = [0] + [f"={cr}/{dr}"] * p["dyears"] + [f"={mr}"] * (5 - p["dyears"])
        for i, f in enumerate(rev_f):
            put(r, 3 + i, f)
        L.reg(SHEET, f"{p['id']}_rev", r, expected=tuple(_rev_flows(p))); rev_row = r; r += 1
        # 成本行
        label_cell(r, f"{p['id']} 成本（交付+运维）", "交付期=合同额×当年成本比例")
        cost_f = [0] + [f"={cr}*${get_column_letter(3 + j)}${cor}" for j in range(p["dyears"])] \
                 + [f"={mr}*{mcr_r}"] * (5 - p["dyears"])
        for i, f in enumerate(cost_f):
            put(r, 3 + i, f)
        L.reg(SHEET, f"{p['id']}_costamt", r, expected=tuple(_cost_flows(p))); cost_row = r; r += 1
        # 净现金流
        label_cell(r, f"{p['id']} 净现金流", "收入−成本", bold=True)
        for i in range(NYRS):
            col = get_column_letter(3 + i)
            put(r, 3 + i, 0 if i == 0 else f"={col}{rev_row}-{col}{cost_row}",
                bold=True, fill=TOTAL_FILL)
        L.reg(SHEET, f"{p['id']}_cf", r, expected=tuple(_flows(p)))
        rows["cf"] = r; r += 1
    r += 1

    # ── 三、指标区：NPV@WACC / IRR（WACC 引 Assumptions）──
    sec("── 三、指标：NPV@WACC 与 IRR（折现率引用 Assumptions!WACC；NPV 折 Y1-5，IRR 用 Y0-5）──")
    label_cell(r, "项目", "备注", bold=True)
    for col, t in [(3, "NPV@WACC（万）"), (4, "IRR")]:
        c = ws.cell(row=r, column=col, value=t)
        c.font = BOLD_WHITE; c.fill = HEADER_FILL; c.alignment = Alignment("center")
    r += 1
    wref = L.ref("Assumptions", "WACC", 0)
    for p in PROJECTS:
        cf_row = prows[p["id"]]["cf"]
        label_cell(r, f"{p['id']} {p['name']}", f"{p['dyears']} 年交付 + 运维")
        put(r, 3, f"=NPV({wref},D{cf_row}:H{cf_row})", font=GREEN_LINK, numFmt=NUM1)
        put(r, 4, f'=IFERROR(IRR(C{cf_row}:H{cf_row}),"—")', font=GREEN_LINK, numFmt=PCT)
        L.reg(SHEET, f"{p['id']}_npv", r, expected=(_npv(_flows(p)),))
        L.reg(SHEET, f"{p['id']}_irr", r)   # IRR 落在 D 列（列偏移 1），期望值单独登记
        L.expected[(SHEET, f"{p['id']}_irr", 1)] = _irr_or_dash(_flows(p))
        r += 1
    r += 1

    # ── 四、敏感性矩阵：P1 NPV = f(合同额档位 × 交付成本比例档位) ──
    p1, rows1 = PROJECTS[0], prows["P1"]
    cr, dr, cor = f"$C${rows1['contract']}", f"$C${rows1['dyears']}", rows1["cost"]
    mr, mcr_r = f"$C${rows1['maint']}", f"$C${rows1['mcr']}"
    sec("── 四、敏感性：P1 NPV（万）＝合同额档位（行）× 交付成本比例档位（列），16 格独立公式重算 ──")
    label_cell(r, "合同额＼交付成本率",
               "收入=合同额×行档；成本=合同额×成本比例×列档；运维不变；按 P1 两年交付结构")
    for j, k in enumerate(MULTS):
        c = ws.cell(row=r, column=3 + j, value=k)
        c.font = BLUE_INPUT; c.number_format = RATIO
        c.fill = HEADER_FILL; c.alignment = Alignment("center")
    L.reg(SHEET, "P1_sens_khdr", r, expected=tuple(MULTS)); hrow = r; r += 1
    for i, m in enumerate(MULTS):
        a = ws.cell(row=r, column=1, value=m); a.font = BLUE_INPUT; a.number_format = RATIO
        exp_row = []
        for j in range(len(MULTS)):
            col = get_column_letter(3 + j)
            f = (f"=NPV({wref},"
                 f"{cr}*$A{r}/{dr}-{cr}*$C${cor}*{col}${hrow},"
                 f"{cr}*$A{r}/{dr}-{cr}*$D${cor}*{col}${hrow},"
                 f"{mr}*(1-{mcr_r}),{mr}*(1-{mcr_r}),{mr}*(1-{mcr_r}))")
            put(r, 3 + j, f, font=GREEN_LINK, numFmt=NUM1)
            exp_row.append(_sens_npv(p1, m, MULTS[j]))
        L.reg(SHEET, f"P1_sens_{i + 1}", r, expected=tuple(exp_row)); r += 1
    r += 1

    # ── 五、决策结论区 ──
    sec("── 五、决策结论 ──")
    order = sorted(PROJECTS, key=lambda p: -_npv(_flows(p)))
    rank_txt = " ＞ ".join(f"{p['id']}（{_npv(_flows(p)):.0f} 万）" for p in order)
    irr1 = _irr_or_dash(_flows(PROJECTS[0]))
    sens_all = [_sens_npv(p1, m, k) for m in MULTS for k in MULTS]
    texts = [
        ("NPV 排名", rank_txt, f"WACC={A.WACC:.0%} 下"),
        ("IRR", f"P1 ≈ {irr1:.0%}（首年仅 −90 万净流出，量级高且口径敏感）；"
                f"P2/P3 全程净流入为正、IRR 无有限解（显示 —），首年即回收", "排序以 NPV 为准"),
        ("敏感性（P1）", f"双向 ±20% 压力下 NPV 区间 {min(sens_all):.0f} ～ {max(sens_all):.0f} 万；"
                        f"最差档（合同额×0.8、成本×1.2）{'仍为正' if min(sens_all) > 0 else '转负——该档位即签约防线，需压成本率或抬合同额'}",
         "16 格公式重算"),
        ("建议", "三项目 NPV 均为正，均可承接；资源受限时优先 P3、P2（交付快、首年回收、运维利润率 70%/65%）；"
                "P1 首年净现金流 −90 万，签约应争取预付款/里程碑与成本发生节奏对齐", "结论基于基准输入"),
    ]
    for label, txt, note in texts:
        label_cell(r, label, note, bold=(label == "建议"))
        c = ws.cell(row=r, column=3, value=txt); c.font = BLACK_FORMULA
        r += 1
    return ws
