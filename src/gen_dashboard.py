"""Dashboard 生成器：FactTable 规范化长表（数据层）+ 四维联动驾驶舱（分析层）。
设计要点（实测 constraints，见 tests/smoke_formulas.py 与本文件自测）：
- FactTable 为引擎影子值直写（黑字数据层，不做逐格影子对比）——长表正确性由
  Dashboard 的 SUMIFS 反查断言保证（反查行 vs Consol_IS!rev 应全 0）；
- 主体/业务线/地区为文本维，"*" 通配在 SUMIFS 里直接可用；年份为数字维，
  "*" 时条件格用 ">0"（"<>" 对数字列匹配不到数字，formulas 引擎实测不支持，
  ">0" 覆盖 2023-2027 全部年份，语义等价"全部期间"）；
- 主体条件格 B3："*"（全部）→"<>合并"。长表同时含『合并行』与『单体行』两套
  粒度，直接用 "*" 会双算；Σ四家单体对外收入 = 合并收入（关联交易全额抵消），
  故"全部"=Σ单体 与 "合并" 两种口径都能与 Consol_IS 勾稽。
"""
import assumptions as A
from style import *
from layout import Layout
from openpyxl.worksheet.datavalidation import DataValidation

NAME_MAP = dict(A.ENTITY_NAMES)
NAME_MAP["CONSOL"] = "合并"


def _entity_name(e):
    return NAME_MAP.get(e, e)


# ─────────────────── FactTable（数据层长表）───────────────────
def gen_facttable(wb, L, eng):
    ws = wb.create_sheet("FactTable")
    ws.sheet_properties.tabColor = TAB["fact"]
    ws["A1"], ws["A1"].font = "FactTable · 规范化长表（数据层）", TITLE
    ws["A2"] = ("Dashboard 取数源：主体 × 业务线 × 地区 × 年份 × 科目 | "
                "引擎影子值直写（黑字）| 万元")
    ws["A2"].font = SUBTITLE
    for j, h in enumerate(["主体", "业务线", "地区", "年份", "科目", "金额"], 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font, c.fill = BOLD_WHITE, HEADER_FILL
        c.alignment = Alignment("center")
    r = 4
    for ent, seg, rg, y, acc, v in eng["facttable"]:
        ws.cell(row=r, column=1, value=_entity_name(ent)).font = BLACK_FORMULA
        ws.cell(row=r, column=2, value=seg).font = BLACK_FORMULA
        ws.cell(row=r, column=3, value=rg).font = BLACK_FORMULA
        c = ws.cell(row=r, column=4, value=int(y))
        c.font, c.number_format = BLACK_FORMULA, "0"
        ws.cell(row=r, column=5, value=acc).font = BLACK_FORMULA
        c = ws.cell(row=r, column=6, value=float(v))
        c.font, c.number_format = BLACK_FORMULA, NUM
        r += 1
    for col, w in zip("ABCDEF", (12, 14, 9, 8, 11, 13)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    return ws


# ─────────────────── Dashboard（分析层）───────────────────
def _dv(ws, options, cell):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"',
                        allow_blank=False, showErrorMessage=True)
    ws.add_data_validation(dv)
    dv.add(cell)


def _year_header(ws, r, label="科目（万元）"):
    c = ws.cell(row=r, column=1, value=label)
    c.font = BOLD
    for i, y in enumerate(A.YEARS):
        c = ws.cell(row=r, column=3 + i, value=y)
        c.number_format = '0"A"' if y <= 2025 else '0"E"'
        c.font, c.fill = BOLD_WHITE, HEADER_FILL
        c.alignment = Alignment("center")


def _sumifs(account):
    """主区统一取数：主体条件 B3 / 业务线 D2 / 地区 F2 / 期间条件 H3"""
    return (f'=SUMIFS(FactTable!$F:$F,FactTable!$A:$A,$B$3,'
            f'FactTable!$B:$B,$D$2,FactTable!$C:$C,$F$2,'
            f'FactTable!$D:$D,$H$3,FactTable!$E:$E,"{account}")')


def gen_dash(wb, L, eng):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = TAB["analysis"]
    ws["A1"], ws["A1"].font = "集团驾驶舱 Dashboard · 四维联动", TITLE
    ws["B1"] = "主体 × 业务线 × 地区 × 期间 下拉联动，SUMIFS 反查 FactTable | 万元"
    ws["B1"].font = SUBTITLE

    # ---- 顶部输入区（黄色输入格 + 数据验证下拉）----
    for col, label in [("A", "主体"), ("C", "业务线"), ("E", "地区"), ("G", "期间")]:
        ws[f"{col}2"], ws[f"{col}2"].font = label, BOLD
    for cell in ("B2", "D2", "F2", "H2"):
        c = ws[cell]
        c.value, c.font, c.fill = "*", BLUE_INPUT, INPUT_FILL
        c.alignment = Alignment("center")
    ws["I2"], ws["I2"].font = "“*”=全部", SUBTITLE
    _dv(ws, ["*", "合并"] + [A.ENTITY_NAMES[e] for e in A.ENTITIES], "B2")
    _dv(ws, ["*"] + A.SEGMENTS, "D2")
    _dv(ws, ["*"] + A.REGIONS, "F2")
    _dv(ws, ["*"] + [str(y) for y in A.YEARS], "H2")

    # ---- 条件格（"*" → 引擎可匹配的判别式）----
    ws["A3"], ws["A3"].font = "取数条件（自动生成）", SUBTITLE
    ws["B3"] = '=IF($B$2="*","<>合并",$B$2)'       # 全部主体 → Σ四家单体（=合并口径）
    ws["G3"], ws["G3"].font = "期间条件", SUBTITLE
    ws["H3"] = '=IF($H$2="*",">0",$H$2)'           # 全部期间 → 所有年份（数字列）
    ws["I3"], ws["I3"].font = "全部主体→<>合并；全部期间→>0", SUBTITLE
    for cell in ("B3", "H3"):
        ws[cell].font, ws[cell].alignment = BLACK_FORMULA, Alignment("center")

    r = 5
    c = ws.cell(row=r, column=1, value="── 维度联动 · 损益视图（当期合计 · 万元）──")
    c.font, c.fill = BOLD, SECTION_FILL
    for col in range(2, 10):
        ws.cell(row=r, column=col).fill = SECTION_FILL
    r += 1
    ws.cell(row=r, column=1, value="科目").font = BOLD
    ws.cell(row=r, column=2, value="口径").font = BOLD
    ws.cell(row=r, column=3, value="金额").font = BOLD
    r += 1

    def mrow(key, label, formula, note=None, bold=False, numFmt=NUM):
        nonlocal r
        line(ws, r, label, formulas=[formula], note=note, bold=bold, numFmt=numFmt, indent=0)
        L.reg("Dashboard", key, r)
        r += 1

    mrow("dash_rev", "营业收入（对外 · 分业务线/地区）", _sumifs("收入"),
         "主体=全部→Σ四家单体对外收入=合并收入")
    mrow("dash_cogs", "营业成本（分业务线）", _sumifs("成本"), "对外成本直加")
    rev_r, cogs_r = r - 2, r - 1
    mrow("dash_gp", "毛利润（维度口径）", f"=C{rev_r}-C{cogs_r}",
         "收入−成本", bold=True)
    gp_r = r - 1
    mrow("dash_rd", "研发费用", _sumifs("研发费用"), "无业务线/地区拆分")
    mrow("dash_sales", "销售费用", _sumifs("销售费用"), "无业务线/地区拆分")
    mrow("dash_gna", "管理费用", _sumifs("管理费用"), "无业务线/地区拆分")
    rd_r, sales_r, gna_r = r - 3, r - 2, r - 1
    mrow("dash_ebit", "EBIT（维度口径）",
         f"=C{gp_r}-C{rd_r}-C{sales_r}-C{gna_r}", "毛利−三费", bold=True)
    mrow("dash_ni", "净利润（全量科目）", _sumifs("净利润"),
         "只随主体/期间联动；主体=全部→Σ单体=合并", bold=True)
    ni_r = r - 1
    mrow("dash_ebitda", "EBITDA（全量科目）", _sumifs("EBITDA"),
         "只随主体/期间联动")
    ebitda_r = r - 1

    # ---- 比率区 ----
    r += 1
    c = ws.cell(row=r, column=1, value="── 比率 ──")
    c.font, c.fill = BOLD, SECTION_FILL
    for col in range(2, 10):
        ws.cell(row=r, column=col).fill = SECTION_FILL
    r += 1
    mrow("dash_gm", "毛利率（维度口径）", f"=IF(C{rev_r}=0,0,C{gp_r}/C{rev_r})", numFmt=PCT)
    mrow("dash_nim", "净利率", f"=IF(C{rev_r}=0,0,C{ni_r}/C{rev_r})", numFmt=PCT)
    mrow("dash_ebitda_m", "EBITDA 率", f"=IF(C{rev_r}=0,0,C{ebitda_r}/C{rev_r})", numFmt=PCT)

    # ---- 维度适用性说明 ----
    r += 1
    ws.cell(row=r, column=1, value=(
        "维度适用性说明：BS/CF 科目（净利润/EBITDA/现金/固定资产）无业务线/地区属性，"
        "本 Dashboard 只做 IS 维度联动，全量科目仅随主体×期间联动；维度口径为对外收入/成本直加，"
        "与合并抵消口径的差异 = 云 T3 内部毛利（0.45×T3，见 Eliminations/IC_Recon）。"
    )).font = SUBTITLE
    for col in range(1, 10):
        ws.cell(row=r, column=col).fill = WARN_FILL
    r += 2

    # ---- ★ 反查断言：FactTable ↔ Consol_IS ----
    c = ws.cell(row=r, column=1, value="── ★ 反查断言 · FactTable ↔ Consol_IS（不受下拉影响）──")
    c.font, c.fill = BOLD, SECTION_FILL
    for col in range(2, 10):
        ws.cell(row=r, column=col).fill = SECTION_FILL
    r += 1
    _year_header(ws, r)
    r += 1
    ft_rev = [f'=SUMIFS(FactTable!$F:$F,FactTable!$A:$A,"合并",FactTable!$B:$B,"*",'
              f'FactTable!$C:$C,"*",FactTable!$D:$D,{y},FactTable!$E:$E,"收入")'
              for y in A.YEARS]
    line(ws, r, "反查 · SUMIFS 全量收入（主体=合并，业务线/地区=*，期间=各年）",
         formulas=ft_rev, note="合并粒度行按业务线×地区重建", indent=0)
    L.reg("Dashboard", "ft_rev_all", r, expected=eng["C"]["IS"]["rev"])
    r += 1
    diff = [f"={L.ref('Dashboard', 'ft_rev_all', i)}-{L.ref('Consol_IS', 'rev', i)}"
            for i in range(5)]
    line(ws, r, "★ 反查差异 vs Consol_IS!rev（应全 0）", formulas=diff,
         bold=True, numFmt=NUM1, fill=TOTAL_FILL, indent=0)
    L.reg("Dashboard", "chk_dash_rev", r)
    r += 2

    # ---- 固定 5 年趋势带（合并口径，跨表直引）----
    c = ws.cell(row=r, column=1, value="── 5 年趋势 · 合并口径（固定，不受下拉影响）──")
    c.font, c.fill = BOLD, SECTION_FILL
    for col in range(2, 10):
        ws.cell(row=r, column=col).fill = SECTION_FILL
    r += 1
    _year_header(ws, r)
    r += 1
    for key, label, sheet, skey, exp in [
            ("trend_rev", "合并营业收入", "Consol_IS", "rev", eng["C"]["IS"]["rev"]),
            ("trend_gp", "合并毛利润", "Consol_IS", "gp", eng["C"]["IS"]["gp"]),
            ("trend_ni", "合并净利润", "Consol_IS", "ni", eng["C"]["IS"]["ni"]),
            ("trend_cash", "期末现金", "Consol_BS", "cash", eng["C"]["BS"]["cash"])]:
        line(ws, r, label, formulas=[f"={L.ref(sheet, skey, i)}" for i in range(5)],
             font=GREEN_LINK, indent=0)
        L.reg("Dashboard", key, r, expected=exp)
        r += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    for i in range(5):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13.5
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 30
    ws.freeze_panes = "A4"
    return ws


def gen(wb, L, eng):
    gen_facttable(wb, L, eng)
    gen_dash(wb, L, eng)
