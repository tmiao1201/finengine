"""统一财务模型样式。FP&A 惯例：蓝字=输入，黑字=公式，负数红色括号。"""
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

FONT = "等线"
BLUE_INPUT = Font(name=FONT, color="0000FF")            # 硬编码输入
BLACK_FORMULA = Font(name=FONT, color="000000")          # 公式
GREEN_LINK = Font(name=FONT, color="007000")             # 跨表链接
BOLD = Font(name=FONT, bold=True)
BOLD_WHITE = Font(name=FONT, bold=True, color="FFFFFF")
TITLE = Font(name=FONT, bold=True, size=14)
SUBTITLE = Font(name=FONT, color="808080", size=9)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")     # 深蓝表头
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")    # 浅蓝分节
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")      # 黄色输入格（下拉）
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")      # 绿色总计
WARN_FILL = PatternFill("solid", fgColor="FCE4EC")       # 红/警示
PASS_FONT = Font(name=FONT, bold=True, color="008000")
FAIL_FONT = Font(name=FONT, bold=True, color="FF0000")

NUM = '#,##0;[Red](#,##0)'
NUM1 = '#,##0.0;[Red](#,##0.0)'
PCT = '0.0%'
RATIO = '0.00"x"'

THIN = Side(style="thin", color="B0B0B0")
MED = Side(style="medium", color="404040")
TOP_BORDER = Border(top=THIN)
TOTAL_BORDER = Border(top=Side(style="thin"), bottom=Side(style="double"))

# Tab 颜色分区
TAB = {"assum": "0070C0", "engine": "808080", "ic": "7030A0", "consol": "00B050",
       "analysis": "ED7D31", "check": "FF0000", "cover": "1F4E79", "fact": "A6A6A6"}

def sheet_scaffold(ws, title, subtitle, years, tab_key, first_data_row=4):
    """统一脚手架：r1 标题 / r3 年份列头 / A 列科目 320 宽 / C-G 数据列 118 宽"""
    ws.sheet_properties.tabColor = TAB[tab_key]
    ws["A1"], ws["A2"] = title, subtitle
    ws["A1"].font = TITLE; ws["A2"].font = SUBTITLE
    ws.cell(row=3, column=1, value="科目（万元）").font = BOLD
    ws.cell(row=3, column=2, value="备注").font = BOLD
    for i, y in enumerate(years):
        c = ws.cell(row=3, column=3 + i, value=y)
        c.number_format = '0"A"' if y <= 2025 else '0"E"'
        c.font = BOLD_WHITE; c.fill = HEADER_FILL; c.alignment = Alignment("center")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    for i in range(len(years)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13.5
    ws.freeze_panes = f"C{first_data_row}"

def line(ws, row, label, values=None, formulas=None, indent=0, bold=False,
         numFmt=NUM, fill=None, note=None, font=None):
    """写一行科目。values=硬输入(蓝)，formulas=公式(黑/绿)。返回 row。"""
    c = ws.cell(row=row, column=1, value=("    " * indent) + label)
    c.font = font or (BOLD if bold else BLACK_FORMULA)
    if fill: c.fill = fill
    if note is not None:
        n = ws.cell(row=row, column=2, value=note); n.font = SUBTITLE
    for i in range(5):
        cell = ws.cell(row=row, column=3 + i)
        if values is not None:
            cell.value = values[i] if i < len(values) else None
            cell.font = BLUE_INPUT
        elif formulas is not None:
            cell.value = formulas[i] if i < len(formulas) else None
            cell.font = font or BLACK_FORMULA
        cell.number_format = numFmt
        if fill: cell.fill = fill
        if bold: cell.border = TOP_BORDER
    return row
