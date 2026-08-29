"""咖啡模型全部 sheet 生成器（单主体精简版）。
复用 finengine 的：style / layout / gen_meta（Meta sheet + meta_sumifs）/ 生成器模式。"""
import meta
from metric_dict import METRICS
from industry_config import SEGMENTS, REGIONS
from style import *
from layout import Layout
from gen_meta import meta_sumifs

Y = meta.YEARS
def m(metric, entity="GROUP", region=None, col="C$3"):
    return meta_sumifs(metric, entity, region, col)

def _scaffold(ws, title, sub, tab):
    sheet_scaffold(ws, title, sub, Y, tab)

# ── Assumptions 指标视图层 ──────────────────────────────
A_MAP = {
    "STORES": ("stores", None), "DAILY_CUPS": ("daily_cups", None), "AVG_TICKET": ("avg_ticket", None),
    "RETAIL_RATIO": ("retail_ratio", None), "COGS_RATE": ("cogs_rate", None),
    "STORE_OPEX_M": ("store_opex_m", None), "MKT_RATE": ("mkt_rate", None),
    "HC_RD": ("headcount_rd", None), "HC_SALES": ("headcount_sales", None), "HC_GNA": ("headcount_gna", None),
    "SAL_RD": ("salary_rd", None), "SAL_SALES": ("salary_sales", None), "SAL_GNA": ("salary_gna", None),
    "CAPEX_STORE": ("capex_per_store", None), "TAX_RATE": ("tax_rate", None),
    "DSO": ("dso_days", None), "DPO": ("dpo_days", None), "DEF_RATE": ("deferred_rate", None),
    "INJECT": ("equity_inject", None), "OPEN_PI": ("open_pi", None),
    "OPEN_RE": ("open_re", None), "OPEN_FA": ("open_store_fa", None),
    "MIX_一线": ("region_mix", "一线"), "MIX_新一线": ("region_mix", "新一线"),
    "MIX_二线及以下": ("region_mix", "二线及以下"), "MIX_海外": ("region_mix", "海外"),
}
def gen_assumptions(wb, L, R):
    ws = wb.create_sheet("Assumptions")
    _scaffold(ws, "山雾咖啡 · 指标视图层", "全部实时取自 Meta（SUMIFS）| 改数请改 Meta!G 列 | 万元", "assum")
    r = 4
    def vrow(key, label, note=None, numFmt=NUM):
        nonlocal r
        metric, rg = A_MAP[key]
        f = [f"={m(metric, region=rg, col=f'{chr(67+i)}$3')}" for i in range(5)]
        exp = [meta.get(metric, i, region=rg) for i in range(5)]
        line(ws, r, label, formulas=f, note=note, numFmt=numFmt, font=GREEN_LINK)
        L.reg("Assumptions", key, r, expected=exp); r += 1
    def sec(t):
        nonlocal r
        c = ws.cell(row=r, column=1, value=t); c.font = BOLD; c.fill = SECTION_FILL; r += 1
    sec("── 量×价 驱动 ──")
    vrow("STORES", "期末门店数（家）", numFmt='#,##0')
    vrow("DAILY_CUPS", "日均杯量/店（杯）", numFmt='#,##0')
    vrow("AVG_TICKET", "客单价（元）", numFmt='0.0')
    vrow("RETAIL_RATIO", "零售占现制比例", numFmt=PCT)
    sec("── 成本 ──")
    vrow("COGS_RATE", "原料成本率", numFmt=PCT)
    vrow("STORE_OPEX_M", "单店月运营（万）", numFmt='0.0')
    vrow("MKT_RATE", "营销费率", numFmt=PCT)
    sec("── 总部与结构 ──")
    for k, lbl, fmt in [("HC_RD", "研发编制", '#,##0'), ("HC_SALES", "销售市场编制", '#,##0'),
                        ("HC_GNA", "管理编制", '#,##0'), ("SAL_RD", "研发人均（万/年）", '0.0'),
                        ("SAL_SALES", "销售人均（万/年）", '0.0'), ("SAL_GNA", "管理人均（万/年）", '0.0'),
                        ("CAPEX_STORE", "单店投入（万）", '0.0'), ("TAX_RATE", "所得税率", PCT),
                        ("DSO", "应收账期（天）", '0'), ("DPO", "应付账期（天）", '0'),
                        ("DEF_RATE", "储值/递延率", PCT), ("INJECT", "当年注资", NUM),
                        ("OPEN_PI", "期初实收资本", NUM), ("OPEN_RE", "期初未分配利润", NUM),
                        ("OPEN_FA", "期初门店资产原值", NUM)]:
        vrow(k, lbl, numFmt=fmt)
    sec("── 收入地区拆分 ──")
    for rg in REGIONS:
        vrow(f"MIX_{rg}", f"{rg} 占比", numFmt=PCT)
    return ws

# ── 门店折旧引擎 ────────────────────────────────────────
def gen_capex(wb, L, R):
    ws = wb.create_sheet("StoreCapex")
    _scaffold(ws, "门店资产 · 批次折旧引擎", "净增门店×单店投入，5 年直线 | 连锁扩张的资产负债表灵魂", "engine")
    ws.column_dimensions["H"].width = 12
    r = 4
    c = ws.cell(row=r, column=1, value="── 批次明细（B=购置年，H=批次原值[取自Meta]）──"); c.font = BOLD; c.fill = SECTION_FILL; r += 1
    rows = []
    stores_prev0 = f"({m('open_store_fa', col='C$3')}/{m('capex_per_store', col='C$3')})"
    for i, y in enumerate(Y):
        ws.cell(row=r, column=1, value=f"{y}批").font = BLACK_FORMULA
        b = ws.cell(row=r, column=2, value=y); b.font = BLUE_INPUT; b.number_format = '0"年"'
        if i == 0:
            h = (f"={m('open_store_fa', col='C$3')}+({m('stores', col='C$3')}-{stores_prev0})"
                 f"*{m('capex_per_store', col='C$3')}")
        else:
            h = (f"=({m('stores', col=f'{chr(67+i)}$3')}-{m('stores', col=f'{chr(66+i)}$3')})"
                 f"*{m('capex_per_store', col=f'{chr(67+i)}$3')}")
        hc = ws.cell(row=r, column=8, value=h); hc.font = GREEN_LINK; hc.number_format = NUM
        for ci in range(5):
            col = get_column_letter(3 + ci)
            cell = ws.cell(row=r, column=3 + ci)
            cell.value = f'=IF(AND({col}$3>=$B{r},{col}$3<$B{r}+5),$H{r}/5,0)'
            cell.number_format = NUM
        rows.append(r); r += 1
    r += 1
    f = [f"=SUM({get_column_letter(3+i)}{rows[0]}:{get_column_letter(3+i)}{rows[-1]})" for i in range(5)]
    line(ws, r, "当年折旧 → COGS", formulas=f, bold=True, note="门店折旧计入营业成本")
    L.reg("StoreCapex", "DEP", r, expected=R["dep"]); dep_row = r; r += 1
    f = [f"=SUMIFS($H${rows[0]}:$H${rows[-1]},$B${rows[0]}:$B${rows[-1]},\"<=\"&{get_column_letter(3+i)}$3)"
         f"-SUM($C${dep_row}:{get_column_letter(3+i)}{dep_row})" for i in range(5)]
    line(ws, r, "门店资产净值 → BS", formulas=f, bold=True)
    L.reg("StoreCapex", "NET", r, expected=R["fa_net"]); net_row = r; r += 1
    # capex 行（供 CF 引用）
    f = [f"={get_column_letter(3+i)}{rows[0]-0}" for i in range(5)]
    line(ws, r, "当年门店 capex → CF",
         formulas=[f"=$H{rows[0]}-{m('open_store_fa', col='C$3')}"] +
                  [f"=$H{rows[i]}" for i in range(1, 5)], bold=True,
         note="2023 扣期初存量=净增")
    L.reg("StoreCapex", "CAPEX", r, expected=R["capex"]); r += 1
    return ws

# ── 三表 ───────────────────────────────────────────────
def aref(L, k, i): return L.ref("Assumptions", k, i)
def cref(L, k, i): return L.ref("StoreCapex", k, i)

def gen_is(wb, L, R):
    ws = wb.create_sheet("IS")
    _scaffold(ws, "山雾咖啡 · 利润表", "量×价：门店×杯量×客单 | 万元", "engine")
    I = R["IS"]; r = 4
    def frow(key, label, formulas, expected, note=None, bold=False, indent=1, fill=None):
        nonlocal r
        line(ws, r, label, formulas=formulas, note=note, bold=bold, indent=indent, fill=fill)
        L.reg("IS", key, r, expected=expected); r += 1
    a = lambda k, i: aref(L, k, i)
    frow("rev_bev", "收入 · 现制饮品",
         [f"={a('STORES',i)}*{a('DAILY_CUPS',i)}*{a('AVG_TICKET',i)}*365/10000" for i in range(5)],
         I["rev_bev"], "门店×杯量×客单×365")
    frow("rev_ret", "收入 · 零售产品",
         [f"={L.ref('IS','rev_bev',i)}*{a('RETAIL_RATIO',i)}" for i in range(5)], I["rev_ret"])
    frow("rev", "营业收入合计",
         [f"={L.ref('IS','rev_bev',i)}+{L.ref('IS','rev_ret',i)}" for i in range(5)], I["rev"], bold=True)
    frow("cogs_mat", "成本 · 原料", [f"={L.ref('IS','rev',i)}*{a('COGS_RATE',i)}" for i in range(5)], I["cogs_mat"])
    frow("cogs_store", "成本 · 门店运营",
         [f"={a('STORES',i)}*{a('STORE_OPEX_M',i)}*12" for i in range(5)], I["cogs_store"], "租金+人工+水电")
    frow("dep", "成本 · 门店折旧", [f"={cref(L,'DEP',i)}" for i in range(5)], I["dep"])
    frow("cogs", "营业成本合计",
         [f"=SUM({L.ref('IS','cogs_mat',i)},{L.ref('IS','cogs_store',i)},{L.ref('IS','dep',i)})" for i in range(5)],
         I["cogs"], bold=True)
    frow("gp", "毛利润",
         [f"={L.ref('IS','rev',i)}-{L.ref('IS','cogs',i)}" for i in range(5)], I["gp"], bold=True)
    frow("mkt", "营销费用", [f"={L.ref('IS','rev',i)}*{a('MKT_RATE',i)}" for i in range(5)], I["mkt"], "补贴退坡")
    frow("rd", "研发费用（数字化）", [f"={a('HC_RD',i)}*{a('SAL_RD',i)}" for i in range(5)], I["rd"])
    frow("sales", "销售费用（总部）", [f"={a('HC_SALES',i)}*{a('SAL_SALES',i)}" for i in range(5)], I["sales"])
    frow("gna", "管理费用", [f"={a('HC_GNA',i)}*{a('SAL_GNA',i)}" for i in range(5)], I["gna"])
    frow("ebit", "EBIT",
         [f"={L.ref('IS','gp',i)}-{L.ref('IS','mkt',i)}-{L.ref('IS','rd',i)}-{L.ref('IS','sales',i)}"
          f"-{L.ref('IS','gna',i)}" for i in range(5)], I["ebit"], bold=True)
    frow("tax", "所得税", [f"=MAX(0,{L.ref('IS','ebit',i)})*{a('TAX_RATE',i)}" for i in range(5)], I["tax"])
    frow("ni", "净利润",
         [f"={L.ref('IS','ebit',i)}-{L.ref('IS','tax',i)}" for i in range(5)], I["ni"], bold=True, fill=TOTAL_FILL)
    return ws

def gen_bs(wb, L, R):
    ws = wb.create_sheet("BS")
    _scaffold(ws, "山雾咖啡 · 资产负债表", "现金=倒挤项 → 结构性平衡 | 万元", "engine")
    B = R["BS"]; r = 4
    def brow(key, label, formulas, expected, note=None, bold=False, indent=1, fill=None):
        nonlocal r
        line(ws, r, label, formulas=formulas, note=note, bold=bold, indent=indent, fill=fill)
        L.reg("BS", key, r, expected=expected); r += 1
    a = lambda k, i: aref(L, k, i)
    i_ = lambda k, i: L.ref("IS", k, i)
    s = lambda k, i: L.ref("BS", k, i)
    p = lambda k, i: s(k, i-1) if i > 0 else "0"
    brow("AR", "应收账款", [f"={i_('rev',i)}*{a('DSO',i)}/365" for i in range(5)], B["AR"], "C端占比高，账期短")
    brow("fa", "门店资产净值", [f"={cref(L,'NET',i)}" for i in range(5)], B["fa"])
    brow("assets_nc", "非现金资产合计", [f"={s('AR',i)}+{s('fa',i)}" for i in range(5)],
         [B["AR"][i]+B["fa"][i] for i in range(5)], bold=True, indent=0)
    brow("AP", "应付账款",
         [f"=({i_('cogs',i)}+{i_('mkt',i)}+{i_('rd',i)}+{i_('sales',i)}+{i_('gna',i)}-{cref(L,'DEP',i)})"
          f"*{a('DPO',i)}/365" for i in range(5)], B["AP"])
    brow("deferred", "递延收益（储值卡）", [f"={i_('rev',i)}*{a('DEF_RATE',i)}" for i in range(5)], B["deferred"])
    brow("liab", "负债合计", [f"={s('AP',i)}+{s('deferred',i)}" for i in range(5)],
         [B["AP"][i]+B["deferred"][i] for i in range(5)], bold=True, indent=0)
    row = r
    brow("paid_in", "实收资本",
         [f"={a('OPEN_PI',0)}+{a('INJECT',0)}"] + [f"={chr(66+i)}{row}+{a('INJECT',i)}" for i in range(1,5)],
         B["paid_in"])
    row = r
    brow("RE", "未分配利润",
         [f"={a('OPEN_RE',0)}+{i_('ni',0)}"] + [f"={chr(66+i)}{row}+{i_('ni',i)}" for i in range(1,5)], B["RE"])
    brow("cash", "货币资金（倒挤）",
         [f"={s('liab',i)}+{s('paid_in',i)}+{s('RE',i)}-{s('assets_nc',i)}" for i in range(5)],
         B["cash"], bold=True, indent=0, fill=TOTAL_FILL)
    brow("assets", "资产总计", [f"={s('cash',i)}+{s('assets_nc',i)}" for i in range(5)],
         [B["cash"][i]+B["AR"][i]+B["fa"][i] for i in range(5)], bold=True, indent=0)
    brow("L_E", "负债+权益合计", [f"={s('liab',i)}+{s('paid_in',i)}+{s('RE',i)}" for i in range(5)],
         [B["AP"][i]+B["deferred"][i]+B["paid_in"][i]+B["RE"][i] for i in range(5)], bold=True, indent=0)
    line(ws, r, "★ BS 平衡差（应全 0）", bold=True); r += 1
    L.reg("BS", "chk_balance", r)
    line(ws, r, "  资产 − 负债 − 权益",
         formulas=[f"={s('assets',i)}-{s('L_E',i)}" for i in range(5)], numFmt=NUM1); r += 1
    return ws

def gen_cf(wb, L, R):
    ws = wb.create_sheet("CF")
    _scaffold(ws, "山雾咖啡 · 现金流量表（间接法）", "与 BS 倒挤现金互为勾稽 | 万元", "engine")
    F = R["CF"]; r = 4
    def crow(key, label, formulas, expected, note=None, bold=False, indent=1):
        nonlocal r
        line(ws, r, label, formulas=formulas, note=note, bold=bold, indent=indent)
        L.reg("CF", key, r, expected=expected); r += 1
    i_ = lambda k, i: L.ref("IS", k, i)
    s = lambda k, i: L.ref("BS", k, i)
    p = lambda k, i: s(k, i-1) if i > 0 else "0"
    c = lambda k, i: L.ref("CF", k, i)
    crow("ni", "净利润", [f"={i_('ni',i)}" for i in range(5)], F["net"][:0] or R["IS"]["ni"])
    crow("da", "加：门店折旧", [f"={cref(L,'DEP',i)}" for i in range(5)], R["dep"])
    crow("d_ar", "减：应收增加", [f"=-({s('AR',i)}-({p('AR',i)}))" for i in range(5)],
         [-(R["BS"]["AR"][i] - (R["BS"]["AR"][i-1] if i else 0)) for i in range(5)])
    crow("d_def", "加：递延增加", [f"={s('deferred',i)}-({p('deferred',i)})" for i in range(5)],
         [R["BS"]["deferred"][i] - (R["BS"]["deferred"][i-1] if i else 0) for i in range(5)])
    crow("d_ap", "加：应付增加", [f"={s('AP',i)}-({p('AP',i)})" for i in range(5)],
         [R["BS"]["AP"][i] - (R["BS"]["AP"][i-1] if i else 0) for i in range(5)])
    crow("cfo", "经营现金流",
         [f"={c('ni',i)}+{c('da',i)}+{c('d_ar',i)}+{c('d_def',i)}+{c('d_ap',i)}" for i in range(5)],
         F["cfo"], bold=True)
    crow("cfi", "投资现金流（开店 capex）", [f"=-{cref(L,'CAPEX',i)}" for i in range(5)], F["cfi"])
    crow("cff", "筹资：股权注资", [f"={aref(L,'INJECT',i)}" for i in range(5)], F["cff"])
    crow("net", "现金净变动",
         [f"={c('cfo',i)}+{c('cfi',i)}+{c('cff',i)}" for i in range(5)], F["net"], bold=True)
    op = meta.get("open_pi", 0) + meta.get("open_re", 0) - meta.get("open_store_fa", 0)
    row = r
    crow("end_cash", "期末现金（CF）",
         [f"={op}+{c('net',0)}"] + [f"={chr(66+i)}{row}+{c('net',i)}" for i in range(1,5)],
         F["end_cash"], bold=True)
    crow("bs_cash", "期末现金（BS 倒挤）", [f"={s('cash',i)}" for i in range(5)], F["bs_cash"])
    line(ws, r, "★ CF=BS 差异（应全 0）", bold=True); r += 1
    L.reg("CF", "chk_tie", r)
    line(ws, r, "  CF推导 − BS倒挤",
         formulas=[f"={c('end_cash',i)}-{c('bs_cash',i)}" for i in range(5)], numFmt=NUM1); r += 1
    return ws

# ── FactTable + Dashboard（三下拉）──────────────────────
def gen_dashboard(wb, L, R):
    from openpyxl.worksheet.datavalidation import DataValidation
    ws = wb.create_sheet("FactTable")
    ws.sheet_properties.tabColor = "A6A6A6"
    for ci, h in enumerate(["entity", "segment", "region", "period", "account", "value"], 1):
        c = ws.cell(row=1, column=ci, value=h); c.font = BOLD_WHITE; c.fill = HEADER_FILL
    for ri, row in enumerate(R["FT"], 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if ci == 6: cell.number_format = NUM
    n = len(R["FT"]) + 1
    for ci, w in enumerate([10, 12, 12, 8, 10, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = "ED7D31"
    ws["A1"] = "山雾咖啡 · 维度看板"; ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["A2"] = "黄格下拉：业务线 / 地区 / 期间（*=全部；期间列数字用 >0 匹配全部）"
    ws["A2"].font = SUBTITLE
    ws.column_dimensions["A"].width = 24
    for col, w in zip("BCDEFG", [13, 13, 13, 13, 13, 13]):
        ws.column_dimensions[col].width = w
    ws["B4"] = "业务线"; ws["B4"].font = BOLD; ws["C4"] = "*"; ws["C4"].fill = INPUT_FILL
    ws["D4"] = "地区"; ws["D4"].font = BOLD; ws["E4"] = "*"; ws["E4"].fill = INPUT_FILL
    ws["F4"] = "期间"; ws["F4"].font = BOLD; ws["G4"] = "*"; ws["G4"].fill = INPUT_FILL
    ws["H4"] = '=IF(G4="*",">0",G4)'; ws["H4"].font = SUBTITLE  # 期间辅助格
    dv1 = DataValidation(type="list", formula1=f'"*,{",".join(SEGMENTS)}"', allow_blank=False)
    dv2 = DataValidation(type="list", formula1=f'"*,{",".join(REGIONS)}"', allow_blank=False)
    dv3 = DataValidation(type="list", formula1='"*,2023,2024,2025,2026,2027"', allow_blank=False)
    ws.add_data_validation(dv1); ws.add_data_validation(dv2); ws.add_data_validation(dv3)
    dv1.add(ws["C4"]); dv2.add(ws["E4"]); dv3.add(ws["G4"])
    labels = ["营业收入", "营业成本", "毛利润", "EBIT", "净利润"]
    r = 6
    for key, lbl in [("收入", labels[0]), ("成本", labels[1]), ("毛利", labels[2]),
                     ("EBIT", labels[3]), ("净利润", labels[4])]:
        ws.cell(row=r, column=1, value=lbl).font = BOLD if key in ("收入", "净利润") else BLACK_FORMULA
        for i in range(5):
            col = get_column_letter(2 + i)
            f = (f"=SUMIFS(FactTable!$F$2:$F${n},FactTable!$E$2:$E${n},\"{key}\","
                 f"FactTable!$B$2:$B${n},$C$4,FactTable!$C$2:$C${n},$E$4,FactTable!$D$2:$D${n},$H$4)")
            cell = ws.cell(row=r, column=2 + i, value=f)
            cell.number_format = NUM
        r += 1
    # 比率区
    ws.cell(row=r, column=1, value="毛利率").font = BLACK_FORMULA
    for i in range(5):
        col = get_column_letter(2 + i)
        c = ws.cell(row=r, column=2 + i, value=f'=IF({col}6=0,0,{col}8/{col}6)')
        c.number_format = PCT
    r += 2
    # 反查断言：全维度收入 vs IS
    ws.cell(row=r, column=1, value="★ 反查：SUMIFS全量 vs IS 收入").font = BOLD; r += 1
    L.reg("Dashboard", "chk_dash_rev", r)
    for i in range(5):
        col = get_column_letter(2 + i)
        f = (f'=SUMIFS(FactTable!$F$2:$F${n},FactTable!$E$2:$E${n},"收入",FactTable!$D$2:$D${n},{i+2023})'
             f"-{L.ref('IS','rev',i)}")
        ws.cell(row=r, column=2 + i, value=f).number_format = NUM1
    return ws

# ── UE 单店模型 ─────────────────────────────────────────
def gen_ue(wb, L, R):
    ws = wb.create_sheet("UE")
    _scaffold(ws, "单位经济 · 单店模型", "volume_price 模式的 UE = 单店 unit economics | 万元", "analysis")
    I = R["IS"]; r = 4
    def urow(key, label, formulas, expected, numFmt=NUM, note=None, bold=False):
        nonlocal r
        line(ws, r, label, formulas=formulas, note=note, bold=bold, numFmt=numFmt)
        L.reg("UE", key, r, expected=expected); r += 1
    a = lambda k, i: aref(L, k, i)
    i_ = lambda k, i: L.ref("IS", k, i)
    urow("rev_store", "单店年收入", [f"={i_('rev',i)}/{a('STORES',i)}" for i in range(5)],
         [I["rev"][i]/meta.get("stores", i) for i in range(5)])
    urow("gp_store", "单店年毛利", [f"=({i_('rev',i)}-{i_('cogs_store',i)}-{i_('cogs_mat',i)})/{a('STORES',i)}"
         for i in range(5)],
         [(I["rev"][i]-I["cogs_store"][i]-I["cogs_mat"][i])/meta.get("stores", i) for i in range(5)],
         note="扣除原料+门店运营（未扣折旧）")
    urow("payback_m", "单店回本期（月）",
         [f"={a('CAPEX_STORE',i)}/({L.ref('UE','gp_store',i)}/12)" for i in range(5)],
         [meta.get("capex_per_store", i)/((I["rev"][i]-I["cogs_store"][i]-I["cogs_mat"][i])
                                          /meta.get("stores", i)/12) for i in range(5)], '0.0"月"',
         "投入45万 ÷ 月毛利", bold=True)
    urow("spx", "同店增速（收入/店）",
         [f"={L.ref('UE','rev_store',i)}/{L.ref('UE','rev_store',i-1)}-1" if i > 0 else "=0"
          for i in range(5)],
         [0.0]+[I["rev"][i]/meta.get("stores", i)/(I["rev"][i-1]/meta.get("stores", i-1))-1 for i in range(1, 5)],
         PCT, "量价拆分的增长质量")
    return ws

# ── Checks + Cover ─────────────────────────────────────
def gen_checks(wb, L, R):
    ws = wb.create_sheet("Checks")
    _scaffold(ws, "勾稽自检面板", "全公式实时链接 | 红灯即不可交付", "check")
    ws.column_dimensions["I"].width = 12
    r = 4
    results = []
    def crow(key, label, formulas, note=None):
        nonlocal r
        line(ws, r, label, formulas=formulas, note=note, numFmt=NUM1)
        st = ws.cell(row=r, column=9,
                     value=f'=IF(AND(ABS(C{r})<0.01,ABS(D{r})<0.01,ABS(E{r})<0.01,ABS(F{r})<0.01,'
                           f'ABS(G{r})<0.01),"✅ PASS","❌ FAIL")')
        st.font = BOLD
        L.reg("Checks", key, r); results.append(r); r += 1
    crow("c1", "C1 BS 平衡（资产−负债−权益）", [f"={L.ref('BS','chk_balance',i)}" for i in range(5)])
    crow("c2", "C2 CF=BS 现金", [f"={L.ref('CF','chk_tie',i)}" for i in range(5)])
    n = len(R["FT"]) + 1
    crow("c3", "C3 FactTable 反查 vs IS 收入",
         [f'=SUMIFS(FactTable!$F$2:$F${n},FactTable!$E$2:$E${n},"收入",FactTable!$D$2:$D${n},{i+2023})'
          f'-{L.ref("IS","rev",i)}' for i in range(5)], "SUMIFS 全量 = IS 公式值")
    crow("c4", "C4 RE 滚动",
         [f"={L.ref('BS','RE',i)}-{L.ref('BS','RE',i-1)}-{L.ref('IS','ni',i)}" if i > 0
          else f"={L.ref('BS','RE',0)}-{meta.get('open_re',0)}-{L.ref('IS','ni',0)}" for i in range(5)])
    r += 1
    c = ws.cell(row=r, column=1, value="总检灯"); c.font = Font(name=FONT, bold=True, size=12)
    lamp = ws.cell(row=r, column=3,
                   value=f'=IF(COUNTIF(I4:I{r-2},"✅ PASS")={len(results)},"🟢 全部通过","🔴 存在失败")')
    lamp.font = Font(name=FONT, bold=True, size=12)
    L.reg("Checks", "lamp", r)
    return ws

def gen_cover(wb, L, R):
    from openpyxl.worksheet.hyperlink import Hyperlink
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_properties.tabColor = TAB["cover"]
    ws["A1"] = "山雾咖啡 · 连锁咖啡财务模型"; ws["A1"].font = Font(name=FONT, bold=True, size=20)
    ws["A2"] = ("finmodel-builder skill 的第二个行业实例（volume_price：门店×杯量×客单）"
                "| 数据与逻辑分离：改数只改 Meta!G | 万元")
    ws["A2"].font = SUBTITLE
    ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 40
    nav = [("Meta", "数据层宽表——唯一取数源"), ("Assumptions", "指标视图层（SUMIFS 取自 Meta）"),
           ("StoreCapex", "门店批次折旧引擎"), ("IS/BS/CF", "三表（倒挤现金勾稽）"),
           ("Dashboard", "三下拉维度看板"), ("UE", "单店单位经济（回本期）"), ("Checks", "勾稽自检")]
    r = 4
    for name, desc in nav:
        t = name.split("/")[0]
        cell = ws.cell(row=r, column=1, value=t)
        cell.hyperlink = Hyperlink(ref=f"A{r}", location=f"'{t}'!A1")
        cell.font = Font(name=FONT, color="0563C1", underline="single")
        ws.cell(row=r, column=2, value=desc).font = BLACK_FORMULA
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="速览").font = BOLD; r += 1
    for i, y in enumerate(Y):
        h = ws.cell(row=r - 1, column=3 + i, value=y)
        h.number_format = '0"A"' if y <= 2025 else '0"E"'
        h.font = BOLD_WHITE; h.fill = HEADER_FILL
    for lbl, key in [("收入", "rev"), ("净利润", "ni"), ("期末现金", "cash")]:
        src = "IS" if key != "cash" else "BS"
        ws.cell(row=r, column=1, value=lbl).font = BLACK_FORMULA
        for i in range(5):
            c = ws.cell(row=r, column=3 + i, value=f"={L.ref(src, key, i)}")
            c.number_format = NUM
        r += 1
    return ws

# ── Calibration 参数校准表（Step 1b 的产物：数据源管道 → 参数修正）────────
def gen_calibration(wb, L, R):
    ws = wb.create_sheet("Calibration")
    _scaffold(ws, "参数校准 · 真实数据源管道", "tsdata·A股连锁餐饮可比（广州酒家/同庆楼/全聚德/佳禾食品 2022-2023 年报）", "assum")
    ws.column_dimensions["B"].width = 30
    r = 4
    hdr = ["参数", "原假设(先验)", "真实区间", "校准值", "来源", "偏差解释"]
    for ci, h in enumerate(hdr, 1):
        c = ws.cell(row=r, column=ci, value=h); c.font = BOLD_WHITE; c.fill = HEADER_FILL
    r += 1
    rows = [
        ("原料成本率 COGS_RATE", "0.33→0.30", "可比毛利率 18-36%（中位18.2%）", "0.33→0.305",
         "tsdata·A股餐饮校准", "现制咖啡外带小店结构优于堂食可比，取区间上部"),
        ("单店月运营 STORE_OPEX_M", "6.8→7.4 万", "人工租金通胀（可比销售费用率中位4.3%）", "6.8→8.0 万",
         "tsdata·A股餐饮校准", "原假设通胀不足，上调"),
        ("营销费率 MKT_RATE", "9%→6%", "可比销售费用率 2.8-10.7%（中位4.3%）", "8.5%→5.2%",
         "tsdata·A股餐饮校准", "咖啡补贴期叙事保留上部，退坡加快"),
        ("净利率（结果校验）", "原模型2027E≈25%", "可比净利率 4-13%（中位7.3%）", "校准后≈22%",
         "对照校验", "仍高于可比：直营咖啡单店模型差异，明示而非隐藏"),
        ("门店数/杯量/客单", "先验假设", "A股无纯咖啡可比（瑞幸在美/粉单）", "保留先验",
         "低置信度假设", "降级链走到可比公司环，参数级差异大，如实标注"),
    ]
    for row in rows:
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = BLACK_FORMULA if ci != 5 else GREEN_LINK
            if ci == 4: c.font = BLUE_INPUT
        r += 1
    ws.cell(row=r+1, column=1, value="管道：tsdata.get_pro().income() → 年报口径毛利率/费用率 → 区间对照 → 参数修正 → Meta source 标注").font = SUBTITLE
    return ws

def gen(wb, L, R):
    gen_assumptions(wb, L, R)
    gen_capex(wb, L, R)
    gen_is(wb, L, R)
    gen_bs(wb, L, R)
    gen_cf(wb, L, R)
    gen_dashboard(wb, L, R)
    gen_ue(wb, L, R)
    gen_checks(wb, L, R)
    gen_calibration(wb, L, R)
    gen_cover(wb, L, R)
