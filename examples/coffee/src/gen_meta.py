"""Meta 宽表生成器 —— 全模型唯一的数据层。
列：A=entity B=segment C=region D=period E=scenario F=metric G=value H=source I=note
改数字只改这里（G 列），全模型联动——这就是 metadata-driven 的日常。"""
from openpyxl.utils import get_column_letter
import meta
from metric_dict import METRICS
from style import *

COLS = ["entity", "segment", "region", "period", "scenario", "metric", "value", "source", "note"]

def meta_sumifs(metric, entity="GROUP", region=None, col="C$3"):
    """从 Meta 表取数的标准 SUMIFS（供其他生成器复用）。"""
    f = (f"SUMIFS(Meta!$G$2:$G$400,Meta!$F$2:$F$400,\"{metric}\","
         f"Meta!$D$2:$D$400,{col}")
    if entity != "*":
        f += f",Meta!$A$2:$A$400,\"{entity}\""
    if region:
        f += f",Meta!$C$2:$C$400,\"{region}\""
    return f + ")"

def gen(wb, L, eng):
    ws = wb.create_sheet("Meta")
    ws.sheet_properties.tabColor = "BF8F00"
    ws["A1"] = "Meta · 数据层宽表（全模型唯一取数源）"
    ws["A1"].font = Font(name=FONT, bold=True, size=13)
    ws["A2"] = ("字段：主体/业务线/地区/期间/场景(A实际·B预算)/指标/值/来源/备注  |  "
                "日常=按期灌数与维护来源；引擎与公式零改动")
    ws["A2"].font = SUBTITLE
    hr = 3
    for ci, name in enumerate(COLS, 1):
        c = ws.cell(row=hr, column=ci, value=name)
        c.font = BOLD_WHITE; c.fill = HEADER_FILL
    widths = [10, 12, 8, 8, 10, 20, 12, 12, 22]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    rows = meta.rows()
    for ri, r in enumerate(rows, hr + 1):
        ws.cell(row=ri, column=1, value=r["entity"])
        ws.cell(row=ri, column=2, value=r["segment"])
        ws.cell(row=ri, column=3, value=r["region"])
        p = ws.cell(row=ri, column=4, value=r["period"])
        p.number_format = '0"A"' if r["scenario"] == "ACTUAL" else '0"E"'
        ws.cell(row=ri, column=5, value=r["scenario"])
        m = ws.cell(row=ri, column=6, value=r["metric"])
        m.font = BLACK_FORMULA
        v = ws.cell(row=ri, column=7, value=r["value"])
        v.number_format = '#,##0.000'
        v.font = BLUE_INPUT  # 数据层：蓝字=可改的数据
        ws.cell(row=ri, column=8, value=r["source"])
        ws.cell(row=ri, column=9, value=r["note"] or
                METRICS.get(r["metric"], {}).get("label", ""))
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{hr}:I{hr + len(rows)}"
    return ws
