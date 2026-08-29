"""Meta 数据层：连锁咖啡实例。接口与 finengine 完全一致（行业无关的契约）：
build_rows / get / bind / load_meta_from_xlsx / rows。"""
import industry_config as _cfg
from metric_dict import METRICS

YEARS = _cfg.YEARS; N = _cfg.N
SEGMENTS = _cfg.SEGMENTS; REGIONS = _cfg.REGIONS

def _scenario(y):
    return "ACTUAL" if y <= 2025 else "BUDGET"

def build_rows():
    rows = []
    def add(metric, value, i, region=None, source="假设初始化"):
        rows.append({"entity": "GROUP", "segment": None, "region": region,
                     "period": YEARS[i], "scenario": _scenario(YEARS[i]),
                     "metric": metric, "value": float(value), "source": source,
                     "note": METRICS.get(metric, {}).get("label", "")})
    for i in range(N):
        add("stores", _cfg.STORES[i], i, source="门店系统")
        add("daily_cups", _cfg.DAILY_CUPS[i], i, source="POS 汇总")
        add("avg_ticket", _cfg.AVG_TICKET[i], i, source="POS 汇总")
        add("retail_ratio", _cfg.RETAIL_RATIO, i)
        add("cogs_rate", _cfg.COGS_RATE[i], i, source="tsdata·A股餐饮校准")
        add("store_opex_m", _cfg.STORE_OPEX_M[i], i, source="tsdata·A股餐饮校准")
        add("mkt_rate", _cfg.MKT_RATE[i], i, source="tsdata·A股餐饮校准")
        for k in ["rd", "sales", "gna"]:
            add(f"headcount_{k}", _cfg.HEADCOUNT[k][i], i)
            add(f"salary_{k}", _cfg.SALARY[k], i)
        add("capex_per_store", _cfg.CAPEX_PER_STORE, i)
        add("tax_rate", _cfg.TAX_RATE, i)
        add("dso_days", _cfg.DSO, i); add("dpo_days", _cfg.DPO, i)
        add("deferred_rate", _cfg.DEFERRED_RATE, i)
        add("equity_inject", _cfg.EQUITY_INJECT[i], i)
        add("open_pi", _cfg.OPEN_PI, i); add("open_re", _cfg.OPEN_RE, i)
        add("open_store_fa", _cfg.OPENING_STORE_FA, i)
        for rg in REGIONS:
            add("region_mix", _cfg.REGION_MIX[rg][i], i, region=rg, source="行业研究")
    return rows

class _Index:
    def __init__(self, rows):
        self.d = {(r["metric"], r["region"], r["period"]): r["value"] for r in rows}
    def get(self, metric, i, region=None):
        k = (metric, region, YEARS[i])
        if k not in self.d:
            raise KeyError(f"Meta 缺行: {k}")
        return self.d[k]

_IDX = _Index(build_rows())

def get(metric, i, entity="GROUP", region=None):
    """引擎统一取数接口（单主体 entity 恒 GROUP，保留参数为的是接口跨项目一致）。"""
    return _IDX.get(metric, i, region)

def bind(rows):
    global _IDX
    _IDX = _Index(rows)

def load_meta_from_xlsx(path):
    from openpyxl import load_workbook
    ws = load_workbook(path)["Meta"]
    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r[3] or not r[5]:
            continue
        rows.append({"entity": r[0], "segment": r[1], "region": r[2], "period": int(r[3]),
                     "scenario": r[4], "metric": r[5], "value": float(r[6]),
                     "source": r[7], "note": r[8] or ""})
    return rows

def rows():
    return build_rows()

if __name__ == "__main__":
    print(f"Meta 宽表：{len(rows())} 行 | 指标字典：{len(METRICS)} 个")
